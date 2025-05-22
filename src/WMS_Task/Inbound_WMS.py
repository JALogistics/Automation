import os
import pandas as pd
from datetime import datetime
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import warnings

# Suppress the specific openpyxl warning
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl.styles.stylesheet')

def get_latest_excel_file(directory_path):
    """
    Get the path of the latest Excel file in the specified directory
    """
    # Get all xlsx files in the directory
    excel_files = [
        os.path.join(directory_path, f) 
        for f in os.listdir(directory_path) 
        if f.endswith('.xlsx')
    ]
    
    if not excel_files:
        raise FileNotFoundError("No Excel files found in the specified directory")
    
    # Get the latest file based on modification time
    latest_file = max(excel_files, key=os.path.getmtime)
    return latest_file

def process_excel_file(file_path):
    """
    Process the Excel file by skipping first two rows and using third row as header
    """
    try:
        # Read Excel file starting from the third row (index 2) as header
        df = pd.read_excel(
            file_path,
            header=2,  # Use the third row as header
            engine='openpyxl'
        )
        
        # Add Ref1 column by concatenating Release number and Container number
        df['Ref1'] = df['Release number'].astype(str) + df['Container number'].astype(str)
        
        return df
    except Exception as e:
        raise Exception(f"Error processing Excel file: {str(e)}")

def save_processed_file(df, output_directory):
    """
    Save the processed DataFrame to the output directory with timestamp and styling
    """
    # Create output directory if it doesn't exist
    os.makedirs(output_directory, exist_ok=True)
    
    # Generate timestamp for the filename
    timestamp = datetime.now().strftime("%Y%m%d")
    output_filename = f"WMS_Stock_Report_{timestamp}.xlsx"
    output_path = os.path.join(output_directory, output_filename)
    
    # Create a writer object with the Excel engine
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Save the DataFrame
        df.to_excel(writer, index=False, sheet_name='WMS_Stock')
        
        # Access the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['WMS_Stock']
        
        # Define styles
        header_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        header_font = Font(bold=True)
        
        # Apply styles to header row
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            
        # Auto-adjust column widths
        for col in range(1, len(df.columns) + 1):
            column_letter = get_column_letter(col)
            worksheet.column_dimensions[column_letter].auto_size = True
    
    return output_path

def main():
    # Directory paths
    input_directory = r"C:\Users\DeepakSureshNidagund\OneDrive - JA Solar GmbH\Documents - Sales Dashboards (BI Solution)\A_WMS Reports\Stock_WMS"
    output_directory = r"C:\Users\DeepakSureshNidagund\OneDrive - JA Solar GmbH\Logistics Reporting\000_Master_Query_Reports\Automation_DB\WMS_Stock_Report"
    
    try:
        # Get the latest Excel file
        latest_file = get_latest_excel_file(input_directory)
        print(f"Processing file: {latest_file}")
        
        # Process the Excel file
        df = process_excel_file(latest_file)
        
        # Save the processed file
        output_file = save_processed_file(df, output_directory)
        print(f"\nFile saved successfully at: {output_file}")
        
        # # Print basic information about the processed data
        # print("\nDataset Info:")
        # print(f"Number of rows: {len(df)}")
        # print(f"Number of columns: {len(df.columns)}")
        # print("\nColumns:", df.columns.tolist())
        
        return df
        
    except Exception as e:
        print(f"Error: {str(e)}")
        return None

if __name__ == "__main__":
    main()
