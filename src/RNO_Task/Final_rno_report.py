import pandas as pd
from datetime import datetime
import os
import shutil
import logging
from openpyxl import load_workbook

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

def transform_dataframe(df):
    """Transform the dataframe according to requirements"""
    try:
        # Drop Rows
        df = df.drop([1])

        # Reset index to make the 1st row (previous 2nd row) as header
        df = df.reset_index(drop=True)
        new_header = df.iloc[0]
        df = df[0:]
        df.columns = new_header
        df = df.drop([0])
        
        # Reset index again after all transformations
        df = df.reset_index(drop=True)
        
        # List of columns to drop
        columns_to_drop = [
            "Qty(PC)", "Qty(plts)", "Power(W)", "Status", "SKU", 
            "Outbound date", "Agreed Delivery date", "Delivery date",
            "On_Sea_Pcs", "In_Stock_Pcs", "Outbounded_Pcs",
            "Outbound_Comparison", "Case1", "Case2", "Case3", "Case4",
            "cnt_Total_Pcs_cdr", "cnt_Outbound_Pcs_cdr",
            "Pcs_from_wms", "New1", "3pls_Planned", "Email_Plan_Outbound"
        ]
        
        # Drop specified columns if they exist
        existing_columns = [col for col in columns_to_drop if col in df.columns]
        if existing_columns:
            df = df.drop(columns=existing_columns)
            logger.info(f"Dropped the following columns: {existing_columns}")

        # Filter out rows where Status Check has specific values or is blank
        if "Status Check " in df.columns:
            initial_rows = len(df)
            # Remove rows where Status Check is F, H, J or blank/null
            df = df[~(df["Status Check "].isin(['F', 'H', 'J']) | 
                     df["Status Check "].isna() | 
                     df["Status Check "].str.strip() == '')]
            rows_removed = initial_rows - len(df)
            logger.info(f"Removed {rows_removed} rows where Status Check was F, H, J, or blank")

        # Filter out rows based on Escalation/Reminder values
        if "Escalation/Reminder" in df.columns:
            initial_rows = len(df)
            df = df[~df["Escalation/Reminder"].isin(["Check", "Status unknown"])]
            rows_removed = initial_rows - len(df)
            logger.info(f"Removed {rows_removed} rows where Escalation/Reminder was 'Check' or 'Status unknown'")
        
        logger.info("Data transformation completed successfully")
        return df
    except Exception as e:
        logger.error(f"Error during data transformation: {str(e)}")
        raise

def copy_rno_report():
    try:
        # Source and destination paths
        source_path = r"C:\Users\DeepakSureshNidagund\OneDrive - JA Solar GmbH\Logistics Reporting\000_Master_Query_Reports\Automation_DB\RNO_Report\RNO_Report.xlsx"
        dest_path = r"C:\Users\DeepakSureshNidagund\OneDrive - JA Solar GmbH\Logistics Reporting\000_Master_Query_Reports\Automation_DB\RNO_Report\Final_RNO_Report.xlsx"
        
        # Read the specific sheet from source Excel file
        logger.info(f"Reading sheet 'RNO Report' from {source_path}")
        df = pd.read_excel(source_path, sheet_name="RNO Report")
        
        # Apply transformations
        logger.info("Applying data transformations...")
        df = transform_dataframe(df)
        
        # Clear existing data in destination workbook while preserving header
        if os.path.exists(dest_path):
            try:
                wb = load_workbook(dest_path)
                if "RNO Report" in wb.sheetnames:
                    sheet = wb["RNO Report"]
                    if sheet.max_row > 1:  # Only delete if there are data rows
                        sheet.delete_rows(2, sheet.max_row - 1)  # Delete all rows except header
                    wb.save(dest_path)
                    wb.close()
                    logger.info("Cleared existing data while preserving header")
            except Exception as e:
                logger.error(f"Error clearing existing data: {str(e)}")
                raise
        
        # Create ExcelWriter object with openpyxl engine
        with pd.ExcelWriter(dest_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # Write the data starting from row 2 (preserving header)
            df.to_excel(writer, sheet_name="RNO Report", index=False, header=False, startrow=1)
        
        logger.info("File processed and saved successfully!")
        return True
        
    except FileNotFoundError:
        logger.error("Source file not found!")
        return False
    except Exception as e:
        logger.error(f"An error occurred: {str(e)}")
        return False

if __name__ == "__main__":
    copy_rno_report()
