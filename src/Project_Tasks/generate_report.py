import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

def generate_project_reports():
    # File paths
    file1_path = r"C:\Users\DeepakSureshNidagund\JA Solar GmbH\Projects - Documents and Tracking\Project Report and Report Genetation\Project_Tracker_List.xlsx"
    file2_dir = r"C:\Users\DeepakSureshNidagund\OneDrive - JA Solar GmbH\Documents - Sales Dashboards (BI Solution)\Z_Factory_Shipment_Report"
    output_dir = r"C:\Users\DeepakSureshNidagund\JA Solar GmbH\Projects - Documents and Tracking\Project Report and Report Genetation\Generated Reports\Pre-files"

    # Ensure output directory exists
    os.makedirs(output_dir, exist_ok=True)

    # Read file1.xlsx
    file1_df = pd.read_excel(file1_path, header=0)

    # Deduplicate based on Internal_Invoice (or entire row if needed)
    file1_df = file1_df.drop_duplicates(subset=["Internal_Invoice"])

    # Filter rows where Status is empty or null
    pending_df = file1_df[file1_df["Status"].isnull() | (file1_df["Status"] == "")]

    # Find the latest file2.xlsx in the directory
    file2_files = [f for f in os.listdir(file2_dir) if f.lower().endswith('.xlsx')]
    if not file2_files:
        raise FileNotFoundError(f"No .xlsx files found in {file2_dir}")
    file2_files.sort(key=lambda x: os.path.getmtime(os.path.join(file2_dir, x)), reverse=True)
    file2_path = os.path.join(file2_dir, file2_files[0])

    # Read file2.xlsx
    file2_df = pd.read_excel(file2_path, header=0)

    # For each row in pending_df, match and save
    for _, row in pending_df.iterrows():
        internal_invoice = row["Internal_Invoice"]
        ref1 = str(row["Ref1"]).strip()
        if not ref1:
            continue  # Skip if Ref1 is empty
        
        # Filter file2_df where 'internal contract' matches internal_invoice
        matched = file2_df[file2_df["internal contract"] == internal_invoice]
        if not matched.empty:
            # Keep only the specified columns
            keep_columns = [
                "external customer full name",
                "ERP number",
                "Internal Delivery Terms or relatedTransactionTerm",
                "internal contract",
                "Eexternal contract",
                "Type",
                "Quantity",
                "Single-chip power",
                "Total wattage",
                "megawatt",
                "BL no.",
                "container no.",
                "Destination Country",
                "Destination Port",
                "Departure Date ( From Port)",
                "Transport Mode",
                "Remark"
            ]
            filtered = matched[[col for col in keep_columns if col in matched.columns]]
            # Map and rename columns for the final report
            column_map = {
                "Name_partner": row.get("3pls", None),
                "Project": row.get("Project_No", None),
                "Customer": "external customer full name",
                "ERP_no_Sales_Invoice": "ERP number",
                "Delivery_Terms": "Internal Delivery Terms or relatedTransactionTerm",
                "Internal_Contract": "internal contract",
                "External_Contract": "external contract",
                "Module-Type": "Type",
                "Container_No": "container no.",
                "Pieces": "Quantity",
                "Wattage": "Single-chip power",
                "Power": "Total wattage",
                "MW": "megawatt",
                "Sea_Freight_Agent": "Sea Freight Agent",
                "B/L_no": "BL no.",
                "Destination_Country": "Destination Country",
                "Destination_Port": "Destination Port",
                "Free_DM_days": "Free DM days",
                "Free_DT_days": "Free DT days",
                "Custom_Clearance_Date": "Custom Clearance Date",
                "Custom_Clearance_No": "Custom Clearance No",
                "Departure_Date ( From Port)": "Departure Date ( From Port)",
                "Transport_Mode": "Transport Mode",
                "Updated_Container No": "Updated Container No",
                "Inbound_Date": "Inbound Date",
                "Planned_Delivery Date": "Planned Delivery Date",
                "Actual_Delivery Date": "Actual Delivery Date",
                "POD sent (Y/N)": "POD sent Y/N",
                "Container_Returned": "Container Returned",
                "Container_Returned date": "Container Returned date",
                "Damgage Claim": "Damgage Container Status",
                "Remark/Comments": "Remark"
            }
            # Prepare the final DataFrame
            final_columns = list(column_map.keys())
            # Add static values from file1 row for Name_partner and Project
            filtered = filtered.rename(columns={v: k for k, v in column_map.items() if v in filtered.columns})
            for col, src in column_map.items():
                if col in ["Name_partner", "Project"]:
                    filtered[col] = column_map[col]
            # Reorder and fill missing columns
            for col in final_columns:
                if col not in filtered.columns:
                    filtered[col] = None
            filtered = filtered[final_columns]
            output_path = os.path.join(output_dir, f"{ref1}.xlsx")
            filtered.to_excel(output_path, index=False)
            # Format only specific header columns: bold and yellow fill
            highlight_columns = [
                "Inbound_Date",
                "Planned_Delivery Date",
                "Actual_Delivery Date",
                "POD sent (Y/N)",
                "Container_Returned",
                "Damgage Container Status"
            ]
            wb = load_workbook(output_path)
            ws = wb.active
            header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            header_font = Font(bold=True)
            for idx, cell in enumerate(ws[1], 1):
                if cell.value in highlight_columns:
                    cell.font = header_font
                    cell.fill = header_fill
            wb.save(output_path)
    print("Project report generation completed.")

if __name__ == "__main__":
    generate_project_reports()
